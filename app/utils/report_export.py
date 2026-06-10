"""Report export helpers: Excel (.xlsx) and PDF via WeasyPrint.

Both functions are pure (no Streamlit dependency) so they are fully testable
in isolation and can be called from any context.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from html import escape

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def build_excel_report(title: str, sections: dict[str, list[dict]]) -> bytes:
    """Build an .xlsx workbook.

    Parameters
    ----------
    title:
        Report title (reserved for future use as a cover sheet).
    sections:
        Maps sheet name -> list of row dicts.  Each dict's keys become column
        headers; every row must share the same keys (missing keys are filled
        with an empty string).

    Returns
    -------
    bytes
        Raw .xlsx bytes suitable for ``st.download_button``.
    """
    wb = openpyxl.Workbook()
    # Remove the default empty sheet created by openpyxl
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    # Cover sheet with report title
    cover = wb.create_sheet("Cover")
    cover["A1"] = title
    cover["A1"].font = Font(bold=True, size=16, color="1F4E79")
    cover["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")

    for sheet_name, rows in sections.items():
        # Sanitise sheet name: max 31 chars, no forbidden chars
        safe_name = "".join(c for c in sheet_name if c.isalnum() or c in " _-")[:31] or "Sheet"
        ws = wb.create_sheet(title=safe_name)

        if not rows:
            ws["A1"] = "(no data)"
            continue

        headers = list(rows[0].keys())

        # Write header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

        # Auto-size columns (approximate: max content width, capped at 60)
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_width = max(
                len(str(header)),
                *(len(str(row.get(header, ""))) for row in rows),
            )
            ws.column_dimensions[col_letter].width = min(max_width + 4, 60)

    if not wb.sheetnames:
        ws = wb.create_sheet("Report")
        ws["A1"] = title

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def build_pdf_report(title: str, sections: dict[str, list[dict]]) -> bytes | None:
    """Build a PDF report via WeasyPrint.

    Parameters
    ----------
    title:
        Report title shown at the top of the document.
    sections:
        Maps section heading -> list of row dicts (same format as
        :func:`build_excel_report`).

    Returns
    -------
    bytes or None
        Raw PDF bytes, or ``None`` if WeasyPrint is not installed.
    """
    try:
        import weasyprint  # noqa: PLC0415
    except (ImportError, OSError):
        return None

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    html_parts: list[str] = [
        "<!DOCTYPE html>",
        "<html><head><meta charset='utf-8'>",
        "<style>",
        "body { font-family: Arial, sans-serif; margin: 40px; color: #1a1a1a; }",
        "h1 { color: #1F4E79; border-bottom: 2px solid #1F4E79; padding-bottom: 8px; }",
        "h2 { color: #2E75B6; margin-top: 32px; }",
        "p.meta { color: #666; font-size: 12px; margin-top: -8px; }",
        "table { border-collapse: collapse; width: 100%; margin-bottom: 24px; font-size: 13px; }",
        "th { background: #1F4E79; color: #fff; padding: 6px 10px; text-align: left; }",
        "td { border: 1px solid #ccc; padding: 5px 10px; }",
        "tr:nth-child(even) td { background: #f4f8fb; }",
        "</style>",
        "</head><body>",
        f"<h1>{escape(title)}</h1>",
        f"<p class='meta'>Generated: {generated_at}</p>",
    ]

    for section_heading, rows in sections.items():
        html_parts.append(f"<h2>{escape(section_heading)}</h2>")
        if not rows:
            html_parts.append("<p><em>(no data)</em></p>")
            continue

        headers = list(rows[0].keys())
        html_parts.append("<table>")
        html_parts.append("<thead><tr>")
        for h in headers:
            html_parts.append(f"<th>{escape(str(h))}</th>")
        html_parts.append("</tr></thead><tbody>")
        for row in rows:
            html_parts.append("<tr>")
            for h in headers:
                html_parts.append(f"<td>{escape(str(row.get(h, '')))}</td>")
            html_parts.append("</tr>")
        html_parts.append("</tbody></table>")

    html_parts.append("</body></html>")
    html_str = "\n".join(html_parts)

    return weasyprint.HTML(string=html_str).write_pdf()
