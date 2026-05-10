"""Excel Integration for DOT Sidewalk Toolkit.

Advanced Excel workbook generation with:
- Pivot table creation from DataFrames
- VLOOKUP-style formula injection
- Conditional formatting (red/yellow/green status)
- Named ranges and data validation
- Multi-sheet workbooks with cross-sheet references
- Dashboard-ready summary sheets with formulas

Requires ``openpyxl``.

Example::

    from socrata_toolkit.excel_integration import (
        ExcelWorkbookBuilder,
        create_pivot_sheet,
        inject_vlookup,
    )

    builder = ExcelWorkbookBuilder()
    builder.add_data_sheet("Inspections", inspections_df)
    builder.add_pivot_sheet("Borough Summary", inspections_df, rows="borough", values="violations", aggfunc="sum")
    builder.add_vlookup_sheet("Lookup", inspections_df, lookup_col="location_id", return_col="status")
    builder.add_dashboard_sheet("Dashboard", metrics)
    builder.save("report.xlsx")
"""

from __future__ import annotations

from typing import Any, Dict, List, Optional, Sequence, Tuple

import pandas as pd


# ---------------------------------------------------------------------------
# Pivot Table Helper
# ---------------------------------------------------------------------------

def create_pivot_table(
    df: pd.DataFrame,
    rows: str | List[str],
    values: str | List[str],
    columns: Optional[str | List[str]] = None,
    aggfunc: str | Dict[str, str] = "sum",
    fill_value: Any = 0,
    margins: bool = True,
    margins_name: str = "Grand Total",
) -> pd.DataFrame:
    """Create a pivot table from a DataFrame.

    Wraps ``pd.pivot_table`` with DOT-friendly defaults including
    Grand Total row/column and zero-fill for missing combinations.

    Args:
        df: Source data.
        rows: Row grouping column(s).
        values: Value column(s) to aggregate.
        columns: Optional column grouping (creates matrix-style pivot).
        aggfunc: Aggregation function(s) -- 'sum', 'mean', 'count', etc.
        fill_value: Value used for missing combinations.
        margins: Include Grand Total row/column.

    Returns:
        Pivoted DataFrame.
    """
    return pd.pivot_table(
        df,
        index=rows,
        values=values,
        columns=columns,
        aggfunc=aggfunc,
        fill_value=fill_value,
        margins=margins,
        margins_name=margins_name,
    ).reset_index()


# ---------------------------------------------------------------------------
# VLOOKUP-style operations
# ---------------------------------------------------------------------------

def vlookup(
    source_df: pd.DataFrame,
    lookup_df: pd.DataFrame,
    source_key: str,
    lookup_key: str,
    return_columns: str | List[str],
    default: Any = None,
) -> pd.DataFrame:
    """Perform a VLOOKUP-style merge, bringing columns from lookup_df into source_df.

    Equivalent to Excel's VLOOKUP: for each row in source_df, find the
    matching row in lookup_df by key and pull in the return columns.

    Args:
        source_df: Primary data (the sheet you're adding columns to).
        lookup_df: Reference table (the lookup range).
        source_key: Key column in source_df.
        lookup_key: Key column in lookup_df.
        return_columns: Column(s) to pull from lookup_df.
        default: Fill value for non-matches.

    Returns:
        source_df with the return columns appended.
    """
    if isinstance(return_columns, str):
        return_columns = [return_columns]

    lookup_subset = lookup_df[[lookup_key] + return_columns].drop_duplicates(subset=[lookup_key])
    merged = source_df.merge(
        lookup_subset,
        left_on=source_key,
        right_on=lookup_key,
        how="left",
        suffixes=("", "_lookup"),
    )
    if lookup_key != source_key and lookup_key in merged.columns:
        merged = merged.drop(columns=[lookup_key])

    if default is not None:
        for col in return_columns:
            if col in merged.columns:
                merged[col] = merged[col].fillna(default)

    return merged


# ---------------------------------------------------------------------------
# INDEX/MATCH equivalent
# ---------------------------------------------------------------------------

def index_match(
    df: pd.DataFrame,
    match_column: str,
    match_value: Any,
    return_column: str,
    default: Any = None,
) -> Any:
    """INDEX/MATCH equivalent: find a value in a column and return from another.

    Args:
        df: The lookup table.
        match_column: Column to search in.
        match_value: Value to search for.
        return_column: Column to return value from.
        default: Value if not found.

    Returns:
        The matched value or default.
    """
    matches = df[df[match_column] == match_value]
    if matches.empty:
        return default
    return matches.iloc[0][return_column]


# ---------------------------------------------------------------------------
# Excel Workbook Builder
# ---------------------------------------------------------------------------

class ExcelWorkbookBuilder:
    """Build complex Excel workbooks with multiple sheets, formulas, and formatting.

    Provides a fluent API for constructing multi-sheet XLSX files with:
    - Data sheets with auto-filter and freeze panes
    - Pivot table sheets
    - VLOOKUP formula sheets
    - Dashboard sheets with conditional formatting
    - Named ranges for cross-sheet references

    Example::

        builder = ExcelWorkbookBuilder()
        builder.add_data_sheet("Raw Data", df)
        builder.add_pivot_sheet("Summary", df, rows="borough", values="violations")
        builder.add_formula_column("Raw Data", "status_lookup",
            '=VLOOKUP(A{row},' "'Ref'!A:B" ',2,FALSE)')
        builder.save("output.xlsx")
    """

    def __init__(self) -> None:
        self._sheets: List[Tuple[str, pd.DataFrame, Dict[str, Any]]] = []
        self._formulas: Dict[str, List[Tuple[str, str]]] = {}  # sheet -> [(col, formula)]
        self._conditional_formats: Dict[str, List[Dict[str, Any]]] = {}

    def add_data_sheet(
        self,
        name: str,
        df: pd.DataFrame,
        freeze_panes: str = "A2",
        auto_filter: bool = True,
    ) -> "ExcelWorkbookBuilder":
        """Add a data sheet with optional freeze panes and auto-filter."""
        self._sheets.append((name, df, {"freeze_panes": freeze_panes, "auto_filter": auto_filter}))
        return self

    def add_pivot_sheet(
        self,
        name: str,
        df: pd.DataFrame,
        rows: str | List[str] = "borough",
        values: str | List[str] = "violations",
        columns: Optional[str | List[str]] = None,
        aggfunc: str = "sum",
    ) -> "ExcelWorkbookBuilder":
        """Add a pivot table sheet."""
        pivot = create_pivot_table(df, rows=rows, values=values, columns=columns, aggfunc=aggfunc)
        self._sheets.append((name, pivot, {"freeze_panes": "A2", "auto_filter": True}))
        return self

    def add_vlookup_sheet(
        self,
        name: str,
        source_df: pd.DataFrame,
        lookup_df: pd.DataFrame,
        source_key: str,
        lookup_key: str,
        return_columns: str | List[str],
    ) -> "ExcelWorkbookBuilder":
        """Add a sheet with VLOOKUP-merged data."""
        merged = vlookup(source_df, lookup_df, source_key, lookup_key, return_columns)
        self._sheets.append((name, merged, {"freeze_panes": "A2", "auto_filter": True}))
        return self

    def add_formula_column(
        self,
        sheet_name: str,
        column_name: str,
        formula_template: str,
    ) -> "ExcelWorkbookBuilder":
        """Register an Excel formula to inject into a column after writing.

        The ``formula_template`` should use ``{row}`` as a placeholder for
        the Excel row number (e.g., ``=A{row}*B{row}``).
        """
        if sheet_name not in self._formulas:
            self._formulas[sheet_name] = []
        self._formulas[sheet_name].append((column_name, formula_template))
        return self

    def add_conditional_format(
        self,
        sheet_name: str,
        column: str,
        rules: List[Dict[str, Any]],
    ) -> "ExcelWorkbookBuilder":
        """Register conditional formatting rules for a column.

        Each rule dict: {"type": "cell", "criteria": ">", "value": 5, "format": {"bg_color": "#FF0000"}}
        Applied via openpyxl after writing.
        """
        if sheet_name not in self._conditional_formats:
            self._conditional_formats[sheet_name] = []
        self._conditional_formats[sheet_name].append({"column": column, "rules": rules})
        return self

    def save(self, path: str) -> str:
        """Write the workbook to disk.

        Returns the output path.
        """
        try:
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback: write without formatting
            pass

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_name, df, opts in self._sheets:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.book[sheet_name]

                if opts.get("freeze_panes"):
                    ws.freeze_panes = opts["freeze_panes"]
                if opts.get("auto_filter"):
                    ws.auto_filter.ref = ws.dimensions

                # Header styling
                try:
                    from openpyxl.styles import PatternFill, Font
                    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                except Exception:
                    pass

            # Inject formulas
            for sheet_name, formula_list in self._formulas.items():
                if sheet_name not in writer.book.sheetnames:
                    continue
                ws = writer.book[sheet_name]
                for col_name, template in formula_list:
                    # Find or add the column
                    col_idx = ws.max_column + 1
                    ws.cell(row=1, column=col_idx, value=col_name)
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=col_idx, value=template.replace("{row}", str(row)))

            # Conditional formatting (simplified: color cells based on rules)
            for sheet_name, cf_list in self._conditional_formats.items():
                if sheet_name not in writer.book.sheetnames:
                    continue
                ws = writer.book[sheet_name]
                try:
                    from openpyxl.styles import PatternFill
                    for cf in cf_list:
                        col = cf["column"]
                        # Find column index
                        col_idx = None
                        for i, cell in enumerate(ws[1], 1):
                            if cell.value == col:
                                col_idx = i
                                break
                        if col_idx is None:
                            continue
                        for rule in cf["rules"]:
                            fill = PatternFill(
                                start_color=rule.get("format", {}).get("bg_color", "FFFFFF"),
                                end_color=rule.get("format", {}).get("bg_color", "FFFFFF"),
                                fill_type="solid",
                            )
                            for row in range(2, ws.max_row + 1):
                                cell = ws.cell(row=row, column=col_idx)
                                try:
                                    val = float(cell.value) if cell.value is not None else 0
                                    criteria = rule.get("criteria", ">")
                                    threshold = float(rule.get("value", 0))
                                    if criteria == ">" and val > threshold:
                                        cell.fill = fill
                                    elif criteria == "<" and val < threshold:
                                        cell.fill = fill
                                    elif criteria == "==" and val == threshold:
                                        cell.fill = fill
                                except (ValueError, TypeError):
                                    pass
                except Exception:
                    pass

        return path

    def sheet_names(self) -> List[str]:
        """Return list of registered sheet names."""
        return [name for name, _, _ in self._sheets]

    def sheet_data(self, name: str) -> Optional[pd.DataFrame]:
        """Return the DataFrame for a given sheet name."""
        for sname, df, _ in self._sheets:
            if sname == name:
                return df
        return None


# ---------------------------------------------------------------------------
# SUMIF / COUNTIF equivalents
# ---------------------------------------------------------------------------

def sumif(df: pd.DataFrame, criteria_col: str, criteria_value: Any, sum_col: str) -> float:
    """Excel SUMIF equivalent: sum values where criteria matches."""
    mask = df[criteria_col] == criteria_value
    return float(df.loc[mask, sum_col].fillna(0).sum())


def countif(df: pd.DataFrame, criteria_col: str, criteria_value: Any) -> int:
    """Excel COUNTIF equivalent: count rows where criteria matches."""
    return int((df[criteria_col] == criteria_value).sum())


def averageif(df: pd.DataFrame, criteria_col: str, criteria_value: Any, avg_col: str) -> float:
    """Excel AVERAGEIF equivalent: average values where criteria matches."""
    mask = df[criteria_col] == criteria_value
    subset = df.loc[mask, avg_col].dropna()
    return float(subset.mean()) if len(subset) > 0 else 0.0
