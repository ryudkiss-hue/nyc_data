"""Comprehensive tests for core.exporters module."""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path
from unittest.mock import patch

import pandas as pd
import pytest

from socrata_toolkit.core.exporters import (
    MongoExporter,
    PostgresExporter,
    XLSXExporter,
)


class TestXLSXExporter:
    """Tests for XLSXExporter class."""

    def test_write_dataframe(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            exporter = XLSXExporter()
            df = pd.DataFrame({
                "id": [1, 2, 3],
                "name": ["Alice", "Bob", "Charlie"],
                "value": [10.5, 20.3, 30.1],
            })
            path = Path(tmpdir) / "test.xlsx"
            exporter.write(df, str(path))

            assert path.exists()
            # Verify we can read it back
            result = pd.read_excel(path, sheet_name="Data")
            assert len(result) == 3
            assert "id" in result.columns
            assert "name" in result.columns

    def test_write_list_of_dicts(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            exporter = XLSXExporter()
            data = [
                {"id": 1, "name": "Alice"},
                {"id": 2, "name": "Bob"},
            ]
            path = Path(tmpdir) / "test.xlsx"
            exporter.write(data, str(path))

            assert path.exists()
            result = pd.read_excel(path, sheet_name="Data")
            assert len(result) == 2

    def test_write_with_freeze_panes(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            exporter = XLSXExporter()
            df = pd.DataFrame({"col": [1, 2, 3]})
            path = Path(tmpdir) / "test.xlsx"
            exporter.write(df, str(path), freeze_panes=True)

            assert path.exists()
            # Load the workbook to verify freeze panes was set
            from openpyxl import load_workbook
            wb = load_workbook(str(path))
            ws = wb["Data"]
            assert ws.freeze_panes == "A2"

    def test_write_with_auto_filter(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            exporter = XLSXExporter()
            df = pd.DataFrame({
                "id": [1, 2, 3],
                "name": ["Alice", "Bob", "Charlie"],
            })
            path = Path(tmpdir) / "test.xlsx"
            exporter.write(df, str(path), auto_filter=True)

            assert path.exists()
            from openpyxl import load_workbook
            wb = load_workbook(str(path))
            ws = wb["Data"]
            # auto_filter.ref should be set
            assert ws.auto_filter.ref is not None

    def test_write_custom_sheet_name(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            exporter = XLSXExporter()
            df = pd.DataFrame({"col": [1, 2]})
            path = Path(tmpdir) / "test.xlsx"
            exporter.write(df, str(path), sheet="CustomSheet")

            result = pd.read_excel(path, sheet_name="CustomSheet")
            assert len(result) == 2

    def test_write_with_empty_dataframe(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            exporter = XLSXExporter()
            df = pd.DataFrame()
            path = Path(tmpdir) / "test.xlsx"
            exporter.write(df, str(path))

            assert path.exists()
            result = pd.read_excel(path, sheet_name="Data")
            assert len(result) == 0

class TestPostgresExporter:
    """Tests for PostgresExporter class (mocked)."""

    def test_init_missing_psycopg(self):
        with patch.dict(sys.modules, {"psycopg": None}):
            with pytest.raises(ImportError, match="postgres extras"):
                PostgresExporter("postgresql://localhost/test")

class TestMongoExporter:
    """Tests for MongoExporter class (mocked)."""

    def test_init_missing_pymongo(self):
        with patch.dict(sys.modules, {"pymongo": None}):
            with pytest.raises(ImportError, match="mongo extras"):
                MongoExporter("mongodb://localhost", "test_db")
