"""Construction List Generator — Phase D anomaly → Excel deliverable."""
from __future__ import annotations

import logging
from dataclasses import dataclass
from typing import Optional

import pandas as pd

logger = logging.getLogger(__name__)

REPAIR_COST_BY_SCOPE: dict[str, dict[str, int]] = {
    "sidewalk_repair": {"low": 800, "high": 3000, "typical": 1500},
    "pedestrian_ramp": {"low": 2000, "high": 8000, "typical": 4500},
    "curb_replacement": {"low": 1500, "high": 5000, "typical": 2800},
    "ada_compliance": {"low": 3000, "high": 12000, "typical": 6000},
    "tree_pit": {"low": 500, "high": 2000, "typical": 900},
}
TIMELINE_WEEKS_BY_SCOPE: dict[str, int] = {
    "sidewalk_repair": 2,
    "pedestrian_ramp": 6,
    "curb_replacement": 4,
    "ada_compliance": 8,
    "tree_pit": 1,
}
_OUTLIER_TO_SCOPE = {
    "critical": "ada_compliance",
    "high": "pedestrian_ramp",
    "moderate": "sidewalk_repair",
    "low": "tree_pit",
}


@dataclass
class ConstructionListConfig:
    """Configuration for ConstructionListGenerator."""

    borough_filter: Optional[str] = None
    min_z_score: float = 2.0
    include_coordinates: bool = True
    color_code_excel: bool = True


class ConstructionListGenerator:
    """Builds construction lists from Phase D MotherDuck anomaly results."""

    def __init__(self, config: Optional[ConstructionListConfig] = None) -> None:
        self.config = config or ConstructionListConfig()

    def build_from_phase_d(self, phase_d_df: pd.DataFrame) -> pd.DataFrame:
        """Transform Phase D anomaly rows into a prioritized construction list.

        Args:
            phase_d_df: DataFrame from v_phase_d_results with columns:
                location_id, borough, latitude, longitude, inspection_count,
                z_score_violations, outlier_class, priority_rank

        Returns:
            DataFrame with construction list including cost and timeline.
        """
        df = phase_d_df.copy()
        if self.config.borough_filter:
            df = df[df["borough"].str.upper() == self.config.borough_filter.upper()]
        df = df[df["z_score_violations"] >= self.config.min_z_score].copy()
        df["scope_category"] = df["outlier_class"].map(_OUTLIER_TO_SCOPE).fillna("sidewalk_repair")
        df["estimated_cost"] = df["scope_category"].map(
            lambda s: REPAIR_COST_BY_SCOPE.get(s, {}).get("typical", 1500)
        )
        df["cost_low"] = df["scope_category"].map(
            lambda s: REPAIR_COST_BY_SCOPE.get(s, {}).get("low", 800)
        )
        df["cost_high"] = df["scope_category"].map(
            lambda s: REPAIR_COST_BY_SCOPE.get(s, {}).get("high", 5000)
        )
        df["timeline_weeks"] = df["scope_category"].map(TIMELINE_WEEKS_BY_SCOPE).fillna(2)
        df = df.sort_values("priority_rank")
        return df.reset_index(drop=True)

    def export_to_excel(self, df: pd.DataFrame, output_path: str) -> None:
        """Write priority-ranked construction list to Excel with color coding.

        Args:
            df: Construction list DataFrame from build_from_phase_d()
            output_path: Path to write Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font
        except ImportError:
            logger.warning("openpyxl not installed, writing plain Excel")
            df.to_excel(output_path, index=False)
            return

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Construction List"

        cols = [
            "priority_rank",
            "location_id",
            "borough",
            "scope_category",
            "estimated_cost",
            "cost_low",
            "cost_high",
            "timeline_weeks",
            "z_score_violations",
            "outlier_class",
            "latitude",
            "longitude",
        ]
        cols = [c for c in cols if c in df.columns]

        header_fill = PatternFill("solid", fgColor="1F4E79")
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)

        COLOR_MAP = {"critical": "FF0000", "high": "FF8C00", "moderate": "FFD700", "low": "90EE90"}
        for ri, row in df[cols].iterrows():
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri + 2, column=ci, value=val)
            if "outlier_class" in cols and self.config.color_code_excel:
                oc = row.get("outlier_class", "")
                color = COLOR_MAP.get(str(oc), "FFFFFF")
                for ci in range(1, len(cols) + 1):
                    ws.cell(row=ri + 2, column=ci).fill = PatternFill("solid", fgColor=color)

        ws2 = wb.create_sheet("Summary")
        ws2.cell(1, 1, "Total Locations")
        ws2.cell(1, 2, len(df))
        ws2.cell(2, 1, "Total Estimated Cost")
        ws2.cell(2, 2, df["estimated_cost"].sum())
        ws2.cell(3, 1, "Cost Range Low")
        ws2.cell(3, 2, df["cost_low"].sum())
        ws2.cell(4, 1, "Cost Range High")
        ws2.cell(4, 2, df["cost_high"].sum())

        wb.save(output_path)
        logger.info("Construction list saved to %s (%d rows)", output_path, len(df))
